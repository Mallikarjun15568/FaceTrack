from . import bp
from flask import jsonify, request, render_template, session, redirect, flash, current_app, make_response
import json
from utils.db import get_db
from datetime import datetime, date, time, timedelta
from blueprints.auth.utils import login_required, role_required
import os

# ==================================================================================
# REPORTS BLUEPRINT - ALL QUERIES USE DATE(check_in_time)
# ==================================================================================
# IMPORTANT: All reports APIs use DATE(check_in_time) to derive attendance dates.
# Legacy 'date' column in attendance table is IGNORED.
# 
# Filtering Rules:
# - WHERE check_in_time IS NOT NULL (excludes legacy/corrupted rows)
# - ORDER BY check_in_time DESC (chronological order)
# 
# This ensures clean, consistent reporting across all endpoints.
# ==================================================================================


# HELPER: Get company settings from database
def get_company_settings():
    """Fetch company name, logo path from settings table"""
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT setting_key, setting_value FROM settings WHERE setting_key IN ('company_name', 'company_logo')")
        rows = cur.fetchall()
        cur.close()
        
        settings = {row['setting_key']: row['setting_value'] for row in rows}
        return {
            'company_name': settings.get('company_name', 'Face Track Attendance System'),
            'company_logo': settings.get('company_logo', '/static/images/logo.png')
        }
    except:
        return {
            'company_name': 'Face Track Attendance System',
            'company_logo': '/static/images/logo.png'
        }


# HELPER: Calculate date ranges based on period type
def get_period_dates(period_type, custom_from=None, custom_to=None):
    """Calculate date range based on period type"""
    today = date.today()
    
    if period_type == "today":
        return today, today
    elif period_type == "last_7_days":
        start = today - timedelta(days=6)
        return start, today
    elif period_type == "this_week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end
    elif period_type == "last_week":
        end = today - timedelta(days=today.weekday() + 1)
        start = end - timedelta(days=6)
        return start, end
    elif period_type == "this_month":
        start = today.replace(day=1)
        if today.month == 12:
            end = today.replace(day=31)
        else:
            end = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        return start, end
    elif period_type == "last_month":
        first_this_month = today.replace(day=1)
        end = first_this_month - timedelta(days=1)
        start = end.replace(day=1)
        return start, end
    elif period_type == "this_quarter":
        quarter = (today.month - 1) // 3
        start = date(today.year, quarter * 3 + 1, 1)
        if quarter == 3:
            end = date(today.year, 12, 31)
        else:
            end = date(today.year, (quarter + 1) * 3 + 1, 1) - timedelta(days=1)
        return start, end
    elif period_type == "this_year":
        return date(today.year, 1, 1), date(today.year, 12, 31)
    elif period_type == "last_year":
        return date(today.year - 1, 1, 1), date(today.year - 1, 12, 31)
    elif period_type == "custom":
        if custom_from and custom_to:
            try:
                start = datetime.strptime(custom_from, "%Y-%m-%d").date()
                end = datetime.strptime(custom_to, "%Y-%m-%d").date()
                return start, end
            except Exception as e:
                print(f"Date parsing error: {e}, using today")
                return today, today
        else:
            # If custom selected but no dates provided, return today
            return today, today
    else:
        return today, today


# HELPER: Calculate summary statistics
def calculate_summary_stats(rows):
    """Calculate attendance statistics from records.

    The helper works off the list of rows returned by the export/query
    functions.  It counts how many rows were marked present, determines
    how many of those were late (check‑in after 09:30) and computes both
    average and total working hours so the various export routines can
    display whatever they need.

    NOTE: this is a *record‑level* summary; absent employees are **not**
    included here.  The dashboard summary API uses a different query
    when employee‑level counts (including absentees) are required.
    """
    if not rows:
        return {
            'total_records': 0,
            'present_count': 0,
            'absent_count': 0,
            'late_count': 0,
            'on_time_count': 0,
            'total_hours': 0,
            'avg_hours': 0,
            'unique_employees': 0
        }
    
    # Treat any row with a check-in/check-out (or common status variants)
    # as a present record. Some rows store status as "check-out" after
    # a completed day, so checking `check_in_time` and status lowercased
    # gives a robust presence detection.
    presence_statuses = {"present", "check-in", "check-out", "checked_in", "checked_out", "checkin", "checkout"}
    present = 0
    for r in rows:
        # Prefer `entry_time` (export alias); fallback to legacy `check_in_time`.
        if r.get('entry_time') or r.get('check_in_time'):
            present += 1
            continue
        st = r.get('status')
        if st and isinstance(st, str) and st.lower() in presence_statuses:
            present += 1
    
    # Count late entries (after 9:30 AM)
    late = 0
    for r in rows:
        # Export queries alias check_in_time as `entry_time`.
        entry = r.get('entry_time') or r.get('check_in_time')
        if entry:
            try:
                if isinstance(entry, datetime):
                    if entry.time() > time(9, 30):
                        late += 1
                else:
                    entry_dt = datetime.strptime(str(entry), "%Y-%m-%d %H:%M:%S")
                    if entry_dt.time() > time(9, 30):
                        late += 1
            except:
                pass
    
    total_hours = 0
    hour_count = 0
    for r in rows:
        wh = r.get('work_hours')
        if wh:
            try:
                if isinstance(wh, timedelta):
                    total_hours += wh.total_seconds() / 3600
                    hour_count += 1
                elif isinstance(wh, str):
                    parts = wh.split(':')
                    if len(parts) >= 2:
                        total_hours += int(parts[0]) + int(parts[1]) / 60
                        hour_count += 1
            except:
                pass
    
    avg_hours = round(total_hours / hour_count, 1) if hour_count > 0 else 0
    unique_employees = len(set(r.get('name', '') for r in rows if r.get('name')))
    
    return {
        'total_records': len(rows),
        'present_count': present,
        'absent_count': 0,  # not derived from raw rows
        'late_count': late,
        'on_time_count': present - late,
        'total_hours': round(total_hours, 1),
        'avg_hours': avg_hours,
        'unique_employees': unique_employees
    }


# SAFE JSON FIX
def jsonify_safe(rows):
    for r in rows:
        for k, v in list(r.items()):
            if isinstance(v, (datetime, date, time, timedelta)):
                r[k] = str(v)
            if isinstance(v, (bytes, bytearray)):
                try:
                    r[k] = v.decode("utf-8")
                except:
                    r[k] = str(v)
    return rows



# ----------------------------------------------------------------------------------
# HELPER: Employee‑level summary (present/late/absent)
# ----------------------------------------------------------------------------------

def get_employee_summary(start_date, end_date, employee_id=None, department=None):
    """Return counts of present, late and absent employees for a given period.

    The logic mirrors ``api_summary`` but accepts optional filters so it
    can be reused by export endpoints.  ``start_date`` and ``end_date``
    are ``date`` objects.
    """
    db = get_db()
    cur = db.cursor(dictionary=True)

    # total active employees in period (respecting filters)
    total_query = [
        "SELECT COUNT(*) AS total FROM employees e WHERE e.status = 'active'",
    ]
    params = []
    # join_date filter
    total_query.append("AND (e.join_date IS NULL OR e.join_date <= %s)")
    params.append(end_date)
    if employee_id:
        total_query.append("AND e.id = %s")
        params.append(employee_id)
    if department:
        total_query.append("AND EXISTS (SELECT 1 FROM departments d WHERE d.id = e.department_id AND d.name = %s)")
        params.append(department)

    cur.execute(" ".join(total_query), tuple(params))
    total = cur.fetchone().get('total', 0) or 0

    # build detailed attendance counts using subquery similar to api_summary
    summary_query = [
        "SELECT",
        "  COUNT(DISTINCT CASE WHEN COALESCE(total_days,0)=0 THEN e.id END) AS absent_count,",
        "  COUNT(DISTINCT CASE WHEN total_days>0 AND COALESCE(late_days,0)=0 THEN e.id END) AS present_ontime_count,",
        "  COUNT(DISTINCT CASE WHEN late_days>0 THEN e.id END) AS late_count,",
        "  COUNT(DISTINCT CASE WHEN total_days>0 THEN e.id END) AS total_present_count",
        "FROM employees e",
        "LEFT JOIN (",
        "    SELECT",
        "        employee_id,",
        "        COUNT(DISTINCT DATE(check_in_time)) AS total_days,",
        "        COUNT(DISTINCT CASE WHEN TIME(check_in_time)>'09:30:00' THEN DATE(check_in_time) END) AS late_days",
        "    FROM attendance",
        "    WHERE DATE(check_in_time) BETWEEN %s AND %s",
        "      AND check_in_time IS NOT NULL",
        "    GROUP BY employee_id",
        ") a ON e.id = a.employee_id",
        "WHERE e.status = 'active'"
    ]
    params = [start_date, end_date]
    if employee_id:
        summary_query.append("AND e.id = %s")
        params.append(employee_id)
    if department:
        summary_query.append("AND EXISTS (SELECT 1 FROM departments d WHERE d.id = e.department_id AND d.name = %s)")
        params.append(department)

    cur.execute(" \n".join(summary_query), tuple(params))
    row = cur.fetchone() or {}
    absent = row.get('absent_count', 0) or 0
    present = row.get('total_present_count', 0) or 0
    late = row.get('late_count', 0) or 0

    percent = 0
    if total > 0:
        percent = round((present / total) * 100)

    cur.close()
    return {
        'total_employees': total,
        'present': present,
        'late': late,
        'absent': absent,
        'attendance_percent': percent
    }


def record_report_history(report_type, report_name, period_type=None, period_from=None, period_to=None, filters=None, generated_by=None, file_size=0):
    """Insert a row into `report_history`. Swallow errors and log them.

    Call after a successful export so the UI can show generated reports.
    """
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("""
            INSERT INTO report_history
            (report_type, report_name, period_type, period_from, period_to, filters, generated_by, file_size)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            report_type,
            report_name,
            period_type,
            period_from if period_from else None,
            period_to if period_to else None,
            json.dumps(filters) if filters else None,
            generated_by or session.get('user_id'),
            file_size
        ))
        db.commit()
        cur.close()
    except Exception as e:
        try:
            current_app.logger.info(f"Could not record report history: {e}")
        except:
            pass





# REPORT HISTORY API
@bp.route("/api/history")
@login_required
@role_required("admin", "hr")
def api_report_history():
    """Return recent generated reports from `report_history` table.

    If the table does not exist or any DB error occurs, return an empty list
    so the frontend does not receive HTML/404 and crash JSON parsing.
    """
    try:
        limit = int(request.args.get('limit', 10))
    except Exception:
        limit = 10

    db = get_db()
    cur = db.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT rh.*, COALESCE(e.full_name, u.username, u.email) AS generated_by_name
            FROM report_history rh
            LEFT JOIN users u ON rh.generated_by = u.id
            LEFT JOIN employees e ON u.id = e.user_id
            ORDER BY rh.generated_at DESC
            LIMIT %s
        """, (limit,))

        rows = cur.fetchall()
        cur.close()

        # Normalize datetime and binary fields
        safe_rows = jsonify_safe(rows)

        # Transform rows to frontend-friendly shape
        transformed = []
        for r in safe_rows:
            item = dict(r)

            # Normalize report_type to uppercase for frontend checks
            rt = (item.get('report_type') or '')
            item['report_type'] = rt.upper() if isinstance(rt, str) else rt

            # Provide from_date/to_date aliases (frontend expects these keys)
            if 'period_from' in item:
                item['from_date'] = item.get('period_from')
            if 'period_to' in item:
                item['to_date'] = item.get('period_to')

            # Provide a short period label
            item['period'] = item.get('period_type') or ''

            # Normalize filters: try to supply employee_name / department_name
            try:
                filters = item.get('filters')
                if isinstance(filters, str):
                    try:
                        filters = json.loads(filters)
                    except:
                        filters = None
                if filters:
                    # If employee id present, try to resolve full name
                    emp_name = None
                    dept_name = None
                    if filters.get('employee'):
                        try:
                            db2 = get_db()
                            cur2 = db2.cursor(dictionary=True)
                            cur2.execute("SELECT full_name FROM employees WHERE id = %s", (filters.get('employee'),))
                            er = cur2.fetchone()
                            cur2.close()
                            if er:
                                emp_name = er.get('full_name')
                        except:
                            emp_name = None
                    if filters.get('department'):
                        # department may already be a name; try to fetch if it's an id
                        dept_val = filters.get('department')
                        if isinstance(dept_val, int) or (isinstance(dept_val, str) and dept_val.isdigit()):
                            try:
                                db3 = get_db()
                                cur3 = db3.cursor(dictionary=True)
                                cur3.execute("SELECT name FROM departments WHERE id = %s", (dept_val,))
                                dr = cur3.fetchone()
                                cur3.close()
                                if dr:
                                    dept_name = dr.get('name')
                                else:
                                    dept_name = str(dept_val)
                            except:
                                dept_name = str(dept_val)
                        else:
                            dept_name = str(dept_val)

                    if emp_name:
                        filters['employee_name'] = emp_name
                    if dept_name:
                        filters['department_name'] = dept_name

                    item['filters'] = filters
            except Exception:
                pass

            transformed.append(item)

        return jsonify({"status": "ok", "history": transformed})
    except Exception as e:
        current_app.logger.info(f"Report history unavailable or empty: {e}")
        try:
            cur.close()
        except:
            pass
        return jsonify({"status": "ok", "history": []})


# PAGE
@bp.route("/")
@login_required
@role_required("admin", "hr")
def reports_page():
    return render_template("reports.html", employee_view=False)


# SUMMARY API (Today's attendance counts)
# Uses DATE(check_in_time) = CURDATE() for filtering
# SUMMARY API (Today's attendance counts)
# Updated to support custom date ranges via period parameter
@bp.route("/api/summary")
@login_required
@role_required("admin", "hr")
def api_summary():
    period_type = request.args.get("period", "today")
    from_date = request.args.get("from")
    to_date = request.args.get("to")
    
    # Calculate date range
    if period_type == "today":
        start_date = date.today()
        end_date = date.today()
    elif period_type == "custom" and from_date and to_date:
        start_date, end_date = get_period_dates(period_type, from_date, to_date)
    else:
        # For all other preset periods, use get_period_dates
        start_date, end_date = get_period_dates(period_type)
    
    db = get_db()
    cur = db.cursor(dictionary=True)

    # ENTERPRISE: Only count employees who had joined by the report period
    # Employees who joined after the period should not be counted as absent
    cur.execute("""
        SELECT COUNT(*) AS total FROM employees 
        WHERE status = 'active' 
        AND (join_date IS NULL OR join_date <= %s)
    """, (end_date,))
    total = cur.fetchone()["total"] or 0

    # Get comprehensive attendance summary for the period
    # Logic: Categorize each employee into one of three groups
    # 1. Never attended = Absent (total_days IS NULL or 0) - ONLY if they had joined
    # 2. Attended but always on-time = Present (On-time)
    # 3. Attended but late at least once = Late
    
    # ENTERPRISE: join_date filter added to prevent counting new employees as absent
    cur.execute("""
        SELECT
            COUNT(DISTINCT CASE 
                WHEN COALESCE(total_days, 0) = 0 
                     AND (e.join_date IS NULL OR e.join_date <= %s) THEN e.id 
            END) as absent_count,
            COUNT(DISTINCT CASE 
                WHEN total_days > 0 AND COALESCE(late_days, 0) = 0 THEN e.id 
            END) as present_ontime_count,
            COUNT(DISTINCT CASE 
                WHEN late_days > 0 THEN e.id 
            END) as late_count,
            COUNT(DISTINCT CASE 
                WHEN total_days > 0 THEN e.id 
            END) as total_present_count
        FROM employees e
        LEFT JOIN (
            SELECT 
                employee_id,
                COUNT(DISTINCT DATE(check_in_time)) as total_days,
                COUNT(DISTINCT CASE 
                    WHEN TIME(check_in_time) > '09:30:00' 
                    THEN DATE(check_in_time) 
                END) as late_days
            FROM attendance
            WHERE DATE(check_in_time) BETWEEN %s AND %s
                AND check_in_time IS NOT NULL
            GROUP BY employee_id
        ) a ON e.id = a.employee_id
        WHERE e.status = 'active'
    """, (end_date, start_date, end_date))
    
    summary = cur.fetchone()
    
    present = summary['present_ontime_count'] or 0
    late = summary['late_count'] or 0
    absent = summary['absent_count'] or 0
    total_attended = summary['total_present_count'] or 0

    cur.close()

    # Attendance percentage based on employees who attended
    percent = 0
    if total > 0:
        percent = round((total_attended / total) * 100)

    return jsonify({
        "status": "ok",
        "summary": {
            "total": total,
            "present": present,
            "late": late,
            "absent": absent,
            "attendance_percent": percent
        }
    })


# DEPARTMENTS
@bp.route("/api/departments")
@login_required
@role_required("admin", "hr")
def api_departments():
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT id, name FROM departments ORDER BY name ASC")
    rows = cur.fetchall()
    cur.close()

    return jsonify({"status": "ok", "departments": rows})


# EMPLOYEES
@bp.route("/api/employees")
@login_required
@role_required("admin", "hr")
def api_employees():
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT e.id, e.full_name, d.name as department
        FROM employees e
        LEFT JOIN departments d ON e.department_id = d.id
        ORDER BY e.full_name ASC
    """)
    rows = cur.fetchall()
    cur.close()

    return jsonify({"status": "ok", "employees": rows})


# TABLE API - Detailed Attendance Records
# Returns all attendance records with DATE(check_in_time) derived dates
# Excludes legacy rows where check_in_time IS NULL
@bp.route("/api/table")
@login_required
@role_required("admin", "hr")
def api_table():
    from_date = request.args.get("from")
    to_date = request.args.get("to")
    employee = request.args.get("employee")
    dept = request.args.get("department")

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = """
        SELECT 
            a.id,
            e.full_name AS name,
            d.name AS department,
            DATE(a.check_in_time) AS date,
            a.check_in_time AS entry_time,
            a.check_out_time AS exit_time,
            a.status,
            (SELECT image_path FROM recognition_logs
             WHERE employee_id = a.employee_id
               AND DATE(timestamp) = DATE(a.check_in_time)
             ORDER BY id DESC LIMIT 1) AS snapshot,
            a.working_hours AS work_hours
        FROM attendance a
        LEFT JOIN employees e ON e.id = a.employee_id
        LEFT JOIN departments d ON d.id = e.department_id
        WHERE a.check_in_time IS NOT NULL
    """

    params = []
    
    # If employee role, automatically filter by their own ID
    if session.get("role") == "employee" and session.get("employee_id"):
        query += " AND a.employee_id = %s"
        params.append(session.get("employee_id"))

    if from_date:
        query += " AND DATE(a.check_in_time) >= %s"
        params.append(from_date)

    if to_date:
        query += " AND DATE(a.check_in_time) <= %s"
        params.append(to_date)

    if employee:
        query += " AND e.full_name = %s"
        params.append(employee)

    if dept:
        query += " AND d.name = %s"
        params.append(dept)

    query += " ORDER BY a.check_in_time DESC"

    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()

    # Normalize path
    for r in rows:
        snap = r.get("snapshot")
        if snap:
            if not snap.startswith("/"):
                if "static/" in snap:
                    r["snapshot"] = "/" + snap[snap.find("static/"):]
                else:
                    r["snapshot"] = "/static/" + snap

    return jsonify({"status": "ok", "records": jsonify_safe(rows)})


# CHART API - Daily & Department-wise attendance trends
# Uses DATE(check_in_time) for date aggregation
# Updated to support custom date ranges via period parameter
@bp.route("/api/chart-data")
@login_required
@role_required("admin", "hr")
def api_chart_data():
    period_type = request.args.get("period", "last_7_days")
    from_date = request.args.get("from")
    to_date = request.args.get("to")
    
    # Calculate date range
    if period_type == "custom" and from_date and to_date:
        start_date_val, end_date_val = get_period_dates(period_type, from_date, to_date)
    else:
        # For all preset periods, use get_period_dates
        start_date_val, end_date_val = get_period_dates(period_type)
    
    db = get_db()
    cur = db.cursor(dictionary=True)

    # Use calculated date range
    start_date = start_date_val
    end_date = end_date_val

    # Get total active employees
    cur.execute("SELECT COUNT(*) AS total FROM employees WHERE status = 'active'")
    total_employees = cur.fetchone()["total"] or 0

    # Get attendance summary for the period (matching summary API logic)
    cur.execute("""
        SELECT
            COUNT(DISTINCT CASE WHEN present_days > 0 AND COALESCE(late_days, 0) = 0 THEN employee_id END) as present_on_time,
            COUNT(DISTINCT CASE WHEN late_days > 0 THEN employee_id END) as present_late,
            COUNT(DISTINCT CASE WHEN COALESCE(present_days, 0) = 0 THEN employee_id END) as absent
        FROM (
            SELECT
                e.id as employee_id,
                COUNT(DISTINCT DATE(a.check_in_time)) as present_days,
                COUNT(DISTINCT CASE WHEN TIME(a.check_in_time) > '09:30:00' THEN DATE(a.check_in_time) END) as late_days
            FROM employees e
            LEFT JOIN attendance a ON e.id = a.employee_id
                AND DATE(a.check_in_time) BETWEEN %s AND %s
                AND a.check_in_time IS NOT NULL
            WHERE e.status = 'active'
            GROUP BY e.id
        ) attendance_summary
    """, (start_date, end_date))

    summary = cur.fetchone()

    # DAILY ATTENDANCE (using selected date range)
    # Present = came on time (before or at 09:30)
    # Late = came late (after 09:30)
    # Absent = total - (present + late)
    cur.execute("""
        SELECT 
            DATE(check_in_time) AS date,
            COUNT(DISTINCT CASE 
                WHEN TIME(check_in_time) <= '09:30:00' 
                THEN employee_id 
            END) AS present,
            COUNT(DISTINCT CASE 
                WHEN TIME(check_in_time) > '09:30:00' 
                THEN employee_id 
            END) AS late,
            0 AS absent  -- Will calculate below
        FROM attendance
        WHERE check_in_time IS NOT NULL
          AND DATE(check_in_time) BETWEEN %s AND %s
        GROUP BY DATE(check_in_time)
        ORDER BY DATE(check_in_time) ASC
    """, (start_date, end_date))
    daily = cur.fetchall()

    # Calculate absent for each day
    for d in daily:
        total_present_today = d["present"] + d["late"]
        d["absent"] = max(0, total_employees - total_present_today)

    # Calculate number of days in selected period
    period_days = (end_date - start_date).days + 1
    
    # DEPARTMENT WISE - Show all departments with attendance percentage for selected period
    # Only count active employees
    cur.execute("""
        SELECT d.name AS department,
               ROUND(
                   (COUNT(a.employee_id) / (COUNT(DISTINCT e.id) * %s)) * 100, 1
               ) AS attendance_percentage
        FROM departments d
        LEFT JOIN employees e ON e.department_id = d.id AND e.status = 'active'
        LEFT JOIN attendance a ON a.employee_id = e.id AND DATE(a.check_in_time) BETWEEN %s AND %s
        GROUP BY d.id, d.name
        ORDER BY attendance_percentage DESC, d.name ASC
    """, (period_days, start_date, end_date))
    departments = cur.fetchall()

    cur.close()

    return jsonify({
        "status": "ok",
        "chart": {
            "daily": jsonify_safe(daily),
            "departments": jsonify_safe(departments),
            "summary": {
                "present": summary['present_on_time'] if summary else 0,
                "late": summary['present_late'] if summary else 2,
                "absent": summary['absent'] if summary else 7
            }
        }
    })


# CSV EXPORT - Simple clean data without formatting (for import/processing)
# Just headers and data - no decorative elements
@bp.route("/api/export/csv")
@login_required
@role_required("admin", "hr")
def export_csv():
    from flask import Response
    import csv
    from io import StringIO

    period_type = request.args.get("period", "custom")
    from_date = request.args.get("from")
    to_date = request.args.get("to")
    employee_id = request.args.get("employee_id")
    employee = request.args.get("user")
    dept = request.args.get("department")

    # Calculate date range - always use get_period_dates for consistency
    start_date, end_date = get_period_dates(period_type, from_date, to_date)
    from_date = str(start_date)
    to_date = str(end_date)

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = """
        SELECT 
            e.id as emp_id,
            e.full_name AS name,
            d.name AS department,
            DATE(a.check_in_time) AS date,
            a.check_in_time AS entry_time,
            a.check_out_time AS exit_time,
            a.status,
            a.working_hours AS work_hours
        FROM attendance a
        LEFT JOIN employees e ON e.id = a.employee_id
        LEFT JOIN departments d ON d.id = e.department_id
        WHERE a.check_in_time IS NOT NULL
    """

    params = []

    if employee_id:
        query += " AND e.id = %s"
        params.append(employee_id)
    elif employee:
        query += " AND e.full_name = %s"
        params.append(employee)

    if from_date:
        query += " AND DATE(a.check_in_time) >= %s"
        params.append(from_date)

    if to_date:
        query += " AND DATE(a.check_in_time) <= %s"
        params.append(to_date)

    if dept:
        query += " AND d.name = %s"
        params.append(dept)

    query += " ORDER BY a.check_in_time DESC"

    cur.execute(query, params)
    rows = cur.fetchall()
    
    # Get company name from settings
    try:
        cur.execute("SELECT setting_value FROM settings WHERE setting_key = 'company_name'")
        company_result = cur.fetchone()
        company_name = company_result['setting_value'] if company_result else 'FaceTrack Pro'
    except:
        company_name = 'FaceTrack Pro'
    
    cur.close()

    # Create CSV with company info and date range at top
    output = StringIO()
    writer = csv.writer(output)

    # compute summary stats from the fetched rows
    stats = calculate_summary_stats(rows)

    # compute employee-level summary (matches dashboard counts)
    emp_summary = get_employee_summary(start_date, end_date, employee_id, dept)

    # Company name header
    writer.writerow([company_name])
    writer.writerow([f"Attendance Report: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"])
    writer.writerow([f"Generated on: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"])
    writer.writerow([])  # Blank line

    # employee-level summary
    writer.writerow(["EMPLOYEE SUMMARY"])
    writer.writerow(["Total Employees", emp_summary['total_employees']])
    writer.writerow(["Present", emp_summary['present']])
    writer.writerow(["Late", emp_summary['late']])
    writer.writerow(["Absent", emp_summary['absent']])
    writer.writerow(["Attendance %", f"{emp_summary['attendance_percent']}%" if emp_summary['total_employees']>0 else "-"])
    writer.writerow([])

    # (Removed record-level summary to keep report concise)

    # Column headers
    writer.writerow(["Date", "Employee Name", "Department", "Status", "Entry Time", "Exit Time", "Working Hours"])

    # Data rows - clean and simple
    for row in rows:
        entry_time = row.get("entry_time", "")
        exit_time = row.get("exit_time", "")
        
        # Format times simply
        entry_str = entry_time.strftime("%Y-%m-%d %H:%M:%S") if isinstance(entry_time, datetime) else str(entry_time) if entry_time else ""
        exit_str = exit_time.strftime("%Y-%m-%d %H:%M:%S") if isinstance(exit_time, datetime) else str(exit_time) if exit_time else ""
        
        writer.writerow([
            str(row.get("date", "")),
            str(row.get("name", "")),
            str(row.get("department", "")),
            str(row.get("status", "")),
            entry_str,
            exit_str,
            str(row.get("work_hours", ""))
        ])

    output.seek(0)

    filename = f"attendance_data_{from_date}_{to_date}.csv" if from_date and to_date else f"attendance_data_{date.today()}.csv"

    # Record history (best-effort)
    try:
        data_bytes = output.getvalue().encode('utf-8')
        file_size = len(data_bytes)
        filters = {"department": dept, "employee": employee or employee_id}
        record_report_history('csv', filename, period_type, from_date, to_date, filters, session.get('user_id'), file_size)
    except:
        pass

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# PDF EXPORT - Enhanced with professional formatting
# Uses DATE(check_in_time) for date column, excludes NULL check_in_time rows
@bp.route("/api/export/pdf")
@login_required
@role_required("admin", "hr")
def export_pdf():
    from flask import make_response
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from io import BytesIO

    period_type = request.args.get("period", "custom")
    from_date = request.args.get("from")
    to_date = request.args.get("to")
    employee_id = request.args.get("employee_id")
    employee = request.args.get("user")
    dept = request.args.get("department")

    # Calculate date range - always use get_period_dates for consistency
    start_date, end_date = get_period_dates(period_type, from_date, to_date)
    from_date = str(start_date)
    to_date = str(end_date)
    
    # Get company settings
    company_settings = get_company_settings()

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = """
        SELECT 
            e.id as emp_id,
            e.full_name AS name,
            d.name AS department,
            DATE(a.check_in_time) AS date,
            a.check_in_time AS entry_time,
            a.check_out_time AS exit_time,
            a.status,
            a.working_hours AS work_hours
        FROM attendance a
        LEFT JOIN employees e ON e.id = a.employee_id
        LEFT JOIN departments d ON d.id = e.department_id
        WHERE a.check_in_time IS NOT NULL
    """

    params = []

    if employee_id:
        query += " AND e.id = %s"
        params.append(employee_id)
    elif employee:
        query += " AND e.full_name = %s"
        params.append(employee)

    if from_date:
        query += " AND DATE(a.check_in_time) >= %s"
        params.append(from_date)

    if to_date:
        query += " AND DATE(a.check_in_time) <= %s"
        params.append(to_date)

    if dept:
        query += " AND d.name = %s"
        params.append(dept)

    query += " ORDER BY a.check_in_time DESC"

    cur.execute(query, params)
    rows = cur.fetchall()
    
    # Calculate summary stats
    stats = calculate_summary_stats(rows)
    # Employee-level summary (matches dashboard counts) used in PDF header/table
    try:
        emp_summary = get_employee_summary(start_date, end_date, employee_id, dept)
    except Exception:
        emp_summary = {'total_employees': 0, 'present': 0, 'late': 0, 'absent': 0, 'attendance_percent': 0}
    
    cur.close()

    # Create PDF with professional formatting
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           topMargin=0.5*inch, bottomMargin=0.5*inch,
                           leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles with professional look
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=4,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        leading=32
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=16,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        leading=20
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=8,
        spaceBefore=16,
        fontName='Helvetica-Bold',
        leading=20
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#374151'),
        spaceAfter=4,
        leading=14
    )

    # Company Header with professional styling
    elements.append(Paragraph(f"<b>{company_settings['company_name']}</b>", title_style))
    elements.append(Paragraph("<font color='#6b7280'>━━━━━━━━━━━━━━━━━━━━━━</font>", subtitle_style))
    elements.append(Paragraph("<b>ATTENDANCE REPORT</b>", subtitle_style))
    elements.append(Spacer(1, 0.3*inch))

    # Professional blue header line
    line_table = Table([['']], colWidths=[10*inch])
    line_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 3, colors.HexColor('#2563eb')),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 0.2*inch))

    # Report Metadata Box with modern styling
    metadata = [
        ['📅 Report Generated:', datetime.now().strftime("%d %B %Y, %I:%M %p")],
        ['📆 Report Period:', f"{from_date if from_date else 'All'} to {to_date if to_date else 'All'}"],
        ['👤 Generated By:', session.get('full_name', 'Admin')]
    ]
    
    if employee or employee_id:
        emp_name = rows[0].get('name') if rows else 'Unknown'
        metadata.append(['👥 Employee:', emp_name])
    else:
        metadata.append(['📊 Scope:', 'All Employees'])
    
    if dept:
        metadata.append(['🏢 Department:', dept])

    meta_table = Table(metadata, colWidths=[2.2*inch, 5.5*inch])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f9ff')),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#2563eb')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 0.3*inch))

    # Summary Statistics Box with gradient effect (employee-level only)
    elements.append(Paragraph("📊 SUMMARY STATISTICS", heading_style))
    elements.append(Spacer(1, 0.1*inch))

    # Keep only employee-level row for a clean professional look
    summary_data = [
        ['👥 Employees', '✅ Present', '⏰ Late', '🚫 Absent', '📈 Attendance %'],
        [
            str(emp_summary.get('total_employees', 0)),
            str(emp_summary.get('present', 0)),
            str(emp_summary.get('late', 0)),
            str(emp_summary.get('absent', 0)),
            f"{emp_summary.get('attendance_percent', 0)}%"
        ]
    ]

    summary_table = Table(summary_data, colWidths=[2*inch]*5)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#e0f2fe')),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e40af')),
        ('INNERGRID', (0, 0), (-1, -1), 1, colors.HexColor('#60a5fa')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.35*inch))

    # Detailed Records Table
    elements.append(Paragraph("📝 DETAILED ATTENDANCE RECORDS", heading_style))
    elements.append(Spacer(1, 0.15*inch))

    # Table data with formatted times
    data = [["📅 Date", "👤 Employee", "🏢 Department", "📊 Status", "🔵 Entry", "🔴 Exit", "⏱️ Hours"]]

    for row in rows:
        entry_time = row.get("entry_time", "")
        exit_time = row.get("exit_time", "")
        
        # Format times
        if isinstance(entry_time, datetime):
            entry_str = entry_time.strftime("%I:%M %p")
        else:
            entry_str = str(entry_time)[:16] if entry_time else "-"
            
        if isinstance(exit_time, datetime):
            exit_str = exit_time.strftime("%I:%M %p")
        else:
            exit_str = str(exit_time)[:16] if exit_time else "-"
        
        data.append([
            str(row.get("date", "")),
            str(row.get("name", ""))[:20],  # Truncate long names
            str(row.get("department", ""))[:14],
            str(row.get("status", "")),
            entry_str,
            exit_str,
            str(row.get("work_hours", ""))[:8]
        ])

    # Create detailed table with adjusted widths
    table = Table(data, colWidths=[1.2*inch, 2*inch, 1.5*inch, 1*inch, 1.2*inch, 1.2*inch, 0.9*inch])
    table.setStyle(TableStyle([
        # Header styling with gradient effect
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Data rows styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        
        # Alternating row colors with better contrast
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        
        # Grid with professional borders
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1f2937')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1f2937')),
        ('INNERGRID', (0, 0), (-1, -1), 0.75, colors.HexColor('#d1d5db')),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.4*inch))

    # Professional Footer
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#6b7280'),
        alignment=TA_CENTER,
        leading=12
    )
    
    # Divider line
    footer_line = Table([['']], colWidths=[10*inch])
    footer_line.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 1, colors.HexColor('#d1d5db')),
    ]))
    elements.append(footer_line)
    elements.append(Spacer(1, 0.1*inch))
    
    # Footer text
    elements.append(Paragraph(f"<i>📄 Computer-generated report • Generated on {datetime.now().strftime('%d %B %Y at %I:%M %p')}</i>", footer_style))
    elements.append(Spacer(1, 0.05*inch))
    elements.append(Paragraph("<b><i>🔒 CONFIDENTIAL - For Internal Use Only</i></b>", footer_style))

    # Build PDF
    doc.build(elements)

    buffer.seek(0)

    filename = f"attendance_report_{from_date}_{to_date}.pdf" if from_date and to_date else f"attendance_report_{date.today()}.pdf"

    response = make_response(buffer.getvalue())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"

    # Record history (best-effort)
    try:
        file_size = len(buffer.getvalue())
        filters = {"department": dept, "employee": employee or employee_id}
        record_report_history('pdf', filename, period_type, from_date, to_date, filters, session.get('user_id'), file_size)
    except:
        pass

    return response


# EXCEL EXPORT - Professional Excel with rich formatting, colors, and multiple sheets
@bp.route("/api/export/excel")
@login_required
@role_required("admin", "hr")
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    period_type = request.args.get("period", "custom")
    from_date = request.args.get("from")
    to_date = request.args.get("to")
    employee_id = request.args.get("employee_id")
    employee = request.args.get("user")
    dept = request.args.get("department")

    # Calculate date range - always use get_period_dates for consistency
    start_date, end_date = get_period_dates(period_type, from_date, to_date)
    from_date = str(start_date)
    to_date = str(end_date)
    
    # Get company settings
    company_settings = get_company_settings()

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = """
        SELECT 
            e.id as emp_id,
            e.full_name AS name,
            d.name AS department,
            DATE(a.check_in_time) AS date,
            a.check_in_time AS entry_time,
            a.check_out_time AS exit_time,
            a.status,
            a.working_hours AS work_hours
        FROM attendance a
        LEFT JOIN employees e ON e.id = a.employee_id
        LEFT JOIN departments d ON d.id = e.department_id
        WHERE a.check_in_time IS NOT NULL
    """

    params = []

    if employee_id:
        query += " AND e.id = %s"
        params.append(employee_id)
    elif employee:
        query += " AND e.full_name = %s"
        params.append(employee)

    if from_date:
        query += " AND DATE(a.check_in_time) >= %s"
        params.append(from_date)

    if to_date:
        query += " AND DATE(a.check_in_time) <= %s"
        params.append(to_date)

    if dept:
        query += " AND d.name = %s"
        params.append(dept)

    query += " ORDER BY a.check_in_time DESC"

    cur.execute(query, params)
    rows = cur.fetchall()
    
    # Calculate summary stats
    stats = calculate_summary_stats(rows)
    
    cur.close()

    # Create Excel workbook
    wb = Workbook()
    
    # ========== SUMMARY SHEET ==========
    ws_summary = wb.active
    ws_summary.title = "📊 Report Summary"
    
    # Define Professional Styles
    title_font = Font(name='Calibri', size=22, bold=True, color='1E3A8A')
    subtitle_font = Font(name='Calibri', size=16, bold=True, color='3B82F6')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    label_font = Font(name='Calibri', size=11, bold=True, color='1F2937')
    value_font = Font(name='Calibri', size=11, color='374151')
    
    # Color Fills
    title_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    blue_header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    light_blue_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    yellow_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    
    # Borders
    thick_border = Border(
        left=Side(style='medium', color='1E40AF'),
        right=Side(style='medium', color='1E40AF'),
        top=Side(style='medium', color='1E40AF'),
        bottom=Side(style='medium', color='1E40AF')
    )
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Alignments
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # ===== COMPANY HEADER =====
    ws_summary['A1'] = f"📋 {company_settings['company_name']}"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].fill = title_fill
    ws_summary['A1'].alignment = center_align
    ws_summary.merge_cells('A1:G1')
    ws_summary.row_dimensions[1].height = 35
    
    ws_summary['A2'] = '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━'
    ws_summary['A2'].font = Font(color='3B82F6')
    ws_summary['A2'].alignment = center_align
    ws_summary.merge_cells('A2:G2')
    
    ws_summary['A3'] = '📊 ATTENDANCE REPORT'
    ws_summary['A3'].font = subtitle_font
    ws_summary['A3'].alignment = center_align
    ws_summary.merge_cells('A3:G3')
    ws_summary.row_dimensions[3].height = 25
    
    # ===== REPORT INFORMATION =====
    row = 5
    ws_summary[f'A{row}'] = '📄 REPORT INFORMATION'
    ws_summary[f'A{row}'].font = Font(size=13, bold=True, color='1E40AF')
    ws_summary.merge_cells(f'A{row}:G{row}')
    ws_summary.row_dimensions[row].height = 22
    row += 1
    
    info_data = [
        ['📅 Report Generated:', datetime.now().strftime("%d %B %Y, %I:%M %p")],
        ['📆 Report Period:', f"{from_date if from_date else 'All Time'} to {to_date if to_date else 'Present'}"],
        ['👤 Generated By:', session.get('full_name', 'Administrator')],
    ]
    
    if employee or employee_id:
        emp_name = rows[0].get('name') if rows else 'Unknown'
        info_data.append(['👥 Employee Filter:', emp_name])
    else:
        info_data.append(['👥 Employee Scope:', 'All Employees'])
    
    if dept:
        info_data.append(['🏢 Department Filter:', dept])
    
    for label, value in info_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = label_font
        ws_summary[f'A{row}'].fill = light_blue_fill
        ws_summary[f'A{row}'].border = thin_border
        ws_summary[f'A{row}'].alignment = left_align
        
        ws_summary[f'B{row}'] = value
        ws_summary[f'B{row}'].font = value_font
        ws_summary[f'B{row}'].border = thin_border
        ws_summary[f'B{row}'].alignment = left_align
        ws_summary.merge_cells(f'B{row}:G{row}')
        ws_summary.row_dimensions[row].height = 20
        row += 1
    
    row += 1
    
    # ===== SUMMARY STATISTICS =====
    ws_summary[f'A{row}'] = '📊 SUMMARY STATISTICS'
    ws_summary[f'A{row}'].font = Font(size=13, bold=True, color='1E40AF')
    ws_summary.merge_cells(f'A{row}:G{row}')
    ws_summary.row_dimensions[row].height = 22
    row += 1
    
    # Stats Table Header (employee-level summary for clarity)
    stats_headers = ['👥 Employees', '✅ Present', '⏰ Late', '🚫 Absent', '📈 Attendance %']
    for col, header in enumerate(stats_headers, start=1):
        cell = ws_summary.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = center_align
        cell.border = thick_border
    ws_summary.row_dimensions[row].height = 25
    row += 1
    
    # Stats Values with color coding (use employee-level summary)
    try:
        emp_summary = get_employee_summary(start_date, end_date, employee_id, dept)
    except Exception:
        emp_summary = {'total_employees': 0, 'present': 0, 'late': 0, 'absent': 0, 'attendance_percent': 0}

    stats_values = [
        emp_summary.get('total_employees', 0),
        emp_summary.get('present', 0),
        emp_summary.get('late', 0),
        emp_summary.get('absent', 0),
        f"{emp_summary.get('attendance_percent', 0)}%"
    ]

    stats_fills = [light_blue_fill, green_fill, yellow_fill, PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'), light_blue_fill]

    for col, (value, fill) in enumerate(zip(stats_values, stats_fills), start=1):
        cell = ws_summary.cell(row=row, column=col)
        cell.value = value
        cell.font = Font(size=13, bold=True, color='1F2937')
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thick_border
    ws_summary.row_dimensions[row].height = 28
    
    # Add note at bottom
    row += 3
    ws_summary[f'A{row}'] = '💡 Tip: Switch to "Attendance Records" sheet for detailed data'
    ws_summary[f'A{row}'].font = Font(size=10, italic=True, color='6B7280')
    ws_summary.merge_cells(f'A{row}:G{row}')
    ws_summary[f'A{row}'].alignment = center_align
    
    # ========== DETAILED RECORDS SHEET ==========
    ws_details = wb.create_sheet(title="📝 Attendance Records")
    
    # Define header fill
    dark_header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    
    # Headers with emojis
    headers = ['📅 Date', '👤 Employee', '🏢 Department', '📊 Status', '🔵 Entry Time', '🔴 Exit Time', '⏱️ Work Hours']
    
    for col, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        cell.fill = dark_header_fill
        cell.alignment = center_align
        cell.border = thick_border
    ws_details.row_dimensions[1].height = 28
    
    # Check if we have data
    if not rows or len(rows) == 0:
        # Add "No data" message
        ws_details['A2'] = 'No attendance records found for the selected criteria'
        ws_details.merge_cells('A2:G2')
        ws_details['A2'].alignment = center_align
        ws_details['A2'].font = Font(size=11, italic=True, color='6B7280')
    else:
        # Data rows with alternating colors and status-based coloring
        for row_idx, row_data in enumerate(rows, start=2):
            entry_time = row_data.get("entry_time", "")
            exit_time = row_data.get("exit_time", "")
            status = str(row_data.get("status", ""))
            
            # Format times nicely
            if isinstance(entry_time, datetime):
                entry_str = entry_time.strftime("%I:%M %p")
            else:
                entry_str = str(entry_time)[:16] if entry_time else "-"
                
            if isinstance(exit_time, datetime):
                exit_str = exit_time.strftime("%I:%M %p")
            else:
                exit_str = str(exit_time)[:16] if exit_time else "-"
            
            data = [
                str(row_data.get("date", "")),
                str(row_data.get("name", "")),
                str(row_data.get("department", "")),
                status,
                entry_str,
                exit_str,
                str(row_data.get("work_hours", ""))
            ]
            
            # Determine row color based on status
            if 'present' in status.lower() or 'on_time' in status.lower():
                row_fill = green_fill if row_idx % 2 == 0 else PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
            elif 'late' in status.lower():
                row_fill = yellow_fill if row_idx % 2 == 0 else PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
            elif 'absent' in status.lower():
                row_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid') if row_idx % 2 == 0 else PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
            else:
                row_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid') if row_idx % 2 == 0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            for col, value in enumerate(data, start=1):
                cell = ws_details.cell(row=row_idx, column=col)
                cell.value = value
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = row_fill
                cell.font = Font(name='Calibri', size=10)
            
            ws_details.row_dimensions[row_idx].height = 20
    
    # Auto-adjust column widths for better readability
    column_widths = {
        'A': 12,  # Date
        'B': 25,  # Employee
        'C': 18,  # Department
        'D': 12,  # Status
        'E': 14,  # Entry Time
        'F': 14,  # Exit Time
        'G': 13,  # Work Hours
    }
    
    for col_letter, width in column_widths.items():
        ws_details.column_dimensions[col_letter].width = width
    
    # Set column widths for summary sheet
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 20
    ws_summary.column_dimensions['E'].width = 20
    ws_summary.column_dimensions['F'].width = 20
    ws_summary.column_dimensions['G'].width = 20
    
    # Freeze header row in details sheet
    ws_details.freeze_panes = 'A2'
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"attendance_report_{from_date}_{to_date}.xlsx" if from_date and to_date else f"attendance_report_{date.today()}.xlsx"
    
    output = make_response(buffer.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename={filename}"
    output.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # Record history (best-effort)
    try:
        file_size = len(buffer.getvalue())
        filters = {"department": dept, "employee": employee or employee_id}
        record_report_history('excel', filename, period_type, from_date, to_date, filters, session.get('user_id'), file_size)
    except:
        pass

    return output

